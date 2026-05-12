from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from flask import make_response
from datetime import datetime


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def make_sheet(ws, headers, rows, col_widths=None):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for col in ws.columns:
            max_length = 0
            for cell in col:
                val = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(val))
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max_length + 2, 50
            )


def build_workbook(sheets_data):
    wb = Workbook()
    first = True
    for entry in sheets_data:
        sheet_name, headers, rows = entry[0], entry[1], entry[2]
        col_widths = entry[3] if len(entry) > 3 else None
        if first:
            ws = wb.active
            ws.title = sheet_name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name[:31])
        make_sheet(ws, headers, rows, col_widths)
    return wb


def workbook_response(wb, filename_prefix):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp.headers["Content-Disposition"] = (
        f"attachment; filename={filename_prefix}_{timestamp}.xlsx"
    )
    return resp
